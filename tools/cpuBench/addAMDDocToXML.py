#!/usr/bin/env python3

from collections import namedtuple
import xml.etree.ElementTree as ET
from xml.dom import minidom
import argparse
import re
from openpyxl import load_workbook

DocEntry = namedtuple('DocEntry', ['mnemonic', 'operands', 'maskingOperand', 'cpuid', 'ops', 'unit', 'lat', 'tp'])

def main():
   parser = argparse.ArgumentParser(description="Add data to XML file from AMD's doc")
   parser.add_argument('-xml')
   parser.add_argument('-xlsx')
   parser.add_argument('-outputXML')
   parser.add_argument('-arch')
   args = parser.parse_args()

   docEntrySet = set()
   mnemonicMap = dict()

   wb = load_workbook(args.xlsx, data_only=True)
   ws = wb['Sheet1'] if (args.arch == 'ZEN5') else wb.active
   for row in ws.iter_rows(min_row=2, values_only=True):
      if row[0] is None: continue
      mnemonicStr = row[0].upper()
      if not mnemonicStr: continue

      if mnemonicStr.endswith('CC'):
         mnemonics = [mnemonicStr.replace('CC', p) for p in ['B', 'BE', 'L', 'LE', 'NB', 'NBE', 'NL', 'NLE', 'NO', 'NP', 'NS', 'NZ', 'O', 'P', 'S', 'Z']]
      elif mnemonicStr in ['INCSSP', 'RDSSP', 'WRSS', 'WRUSS']:
         mnemonics = [f'{mnemonicStr}D', f'{mnemonicStr}Q']
      elif mnemonicStr == 'RETF':
         mnemonics = ['RETFD', 'RETFW', 'RETFQ']
      elif mnemonicStr == 'SYSRET':
         mnemonics = ['SYSRETQ']
      else:
         mnemonics = mnemonicStr.replace(' (NEAR)', '').replace(' NEAR', '').replace('CC', '').strip().split('/')

      for mnemonic in mnemonics:
         if mnemonic in ['AAA', 'AAD', 'AAM', 'AAS', 'ARPL', 'BOUND', 'DAA', 'DAS', 'INTO', 'JCXZ', 'LDS', 'LES','POPA', 'POPAD', 'POPD', 'POPFD', 'PUSHA', 'PUSHAD', 'PUSHFD']:
            # 32-bit instructions
            continue
         if mnemonic in ['CMPS', 'FCLEX', 'FINIT', 'FSAVE', 'FSTCW', 'FSTENV', 'FSTSW', 'INS', 'LODS', 'LOOPNZ', 'LOOPZ', 'MOVS', 'OUTS', 'PCLMULHQHQDQ', 'PCLMULHQLQDQ', 'PCLMULLQHQDQ', 'PCLMULLQLQDQ', 'RDPRU', 'SAL', 'SCAS', 'STOS', 'VGATHERDD', 'VGATHERDQ', 'VGATHERQD', 'VGATHERQQ','VPCLMULHQHQDQ', 'VPCLMULHQLQDQ', 'VPCLMULLQHQDQ', 'VPCLMULLQLQDQ', 'WAIT', 'XLATB']:
            # missing in XED
            continue
         if mnemonic in ['INT1', 'JECXZ']:
            # missing from XML file
            continue
         if mnemonic in ['VFMADD213SS', 'VFNMADD132SD', 'VFMSUB231SS', 'VFMADD231SS', 'VFMADD213SD', 'VFNMSUB231SD', 'VFNMADD213SD', 'VFNMADD231SD', 'VFNMSUB132SS', 'VFMSUB132SD', 'VFMADD132SD', 'VFMADD132SS', 'VFNMSUB132SD', 'VFNMADD213SS', 'VFMSUB213SD', 'VFNMADD132SS', 'VFNMSUB231SS', 'VFNMSUB213SS', 'VFMADD231SD', 'VFNMSUB213SD', 'VFMSUB213SS', 'VFMSUB231SD', 'VFNMADD231SS', 'VFMSUB132SS'] and row[1] == 'ymm1':
            # no ymm variant of this instruction
            continue

         cpuidCol = 6 if (args.arch in ['ZEN+', 'ZEN2', 'ZEN3']) else 7 if (args.arch == 'ZEN4') else 8
         cpuid = (row[cpuidCol] or '').upper()

         operands = row[2:6] if (args.arch in ['ZEN5']) else row[1:5]
         operands = tuple(op.strip() for op in operands if (op is not None))

         maskingOperand = '' if (args.arch in ['ZEN+', 'ZEN2', 'ZEN3']) else (row[cpuidCol-2] or '')

         ops = row[cpuidCol+1]
         if ops == 'not supported':
            continue

         unit = row[cpuidCol+2]
         lat = row[cpuidCol+3]
         tp = row[cpuidCol+4]

         if (ops is None) and (unit is None) and (lat is None) and (tp is None):
            continue

         de = DocEntry(mnemonic, operands, maskingOperand, cpuid, ops, unit, lat, tp)
         docEntrySet.add(de)
         mnemonicMap.setdefault(mnemonic, []).append(de)

   iclassAsmDict = dict()

   root = ET.parse(args.xml).getroot()
   for instrNode in root.iter('instruction'):
      if 'APX' in instrNode.attrib.get('extension', ''):
         continue
      iclass = instrNode.attrib['iclass']
      asm = instrNode.attrib['asm']
      iclassAsmDict.setdefault(iclass, set()).add(instrNode)
      iclassAsmDict.setdefault(re.sub('{.*} ', '', asm), set()).add(instrNode)

   #for x in set(op for de in docList for op in de.operands):
   #   print(x)

   xmlToDocDict = dict()

   for de in sorted(docEntrySet, key=str):
      if de.mnemonic not in iclassAsmDict:
         print('no XML entry found for ' + str(de))
         continue
      elif len(iclassAsmDict[de.mnemonic]) == 1:
         xmlToDocDict[next(iter(iclassAsmDict[de.mnemonic]))] = de
         continue

      xmlFound = False
      for instrNode in iclassAsmDict[de.mnemonic]:
         if ('AVX512' in de.cpuid) != ('AVX512' in instrNode.attrib['extension']):
            continue

         explXmlOperands = [op for op in instrNode.findall('./operand') if not op.attrib.get('suppressed', '') == '1'
                            and not op.attrib.get('implicit', '') == '1' and not op.attrib.get('opmask', '') == '1']
         xmlZeroing = (instrNode.attrib.get('zeroing', '') == '1')
         docOperands = [op for op in de.operands if op not in [None, 'v']]

         if (not docOperands and any(op.attrib['type'] == 'mem' for op in explXmlOperands) and
             any(len(instrNode2.findall('./operand[@type="mem"]')) == 0 for instrNode2 in iclassAsmDict[de.mnemonic] if instrNode != instrNode2)):
            continue

         if docOperands and explXmlOperands and (len(explXmlOperands) != len(docOperands)):
            if any(len(explXmlOperands) == len([op for op in de2.operands if op is not None]) for de2 in mnemonicMap[de.mnemonic] if de!=de2):
               continue

         if any(op.text == 'CL' for op in instrNode.findall('./operand')) != any('CL' in op.upper() for op in docOperands):
            continue

         if docOperands and explXmlOperands:
            xmlOperands = explXmlOperands
         else:
            xmlOperands = [op for op in instrNode.findall('./operand')]

         invalid = False
         for docOp, xmlOp in zip(docOperands, xmlOperands):
            if xmlOp.attrib['type'] == 'mem' and set(de.operands) == {None}:
               invalid = True
               break
            if docOp is None: continue
            if docOp in ['pntr16/mem16:16/32']: continue

            if xmlOp.attrib['type'] == 'reg':
               if (docOp == 'segmentReg') or ('SS' in docOp.upper()):
                  if xmlOp.attrib.get('implicit', '') == '1': continue
               elif xmlOp.text.startswith('K'):
                  if docOp.startswith('k'): continue
               elif 'MM' in xmlOp.text:
                  if (('mmx' in docOp) or ('M' in docOp)) and xmlOp.text.startswith('MM'): continue
                  if (not 'mmx' in docOp) and (not 'AX' in docOp):
                     if 'x' in docOp and xmlOp.text.startswith('XMM'): continue
                     if 'y' in docOp and xmlOp.text.startswith('YMM'): continue
                     if 'z' in docOp and xmlOp.text.startswith('ZMM'): continue
               elif docOp in ['reg', 'reg/mem', 'reg/imm', 'reg/mem', 'reg/mem/imm', 'r', 'r/i', 'r/m', 'i/r', 'i/r/m', 'g', 'g/m', 'gx/m'] and xmlOp.attrib.get('implicit', '') != '1':
                  continue
               elif (docOp in ['Sti', 's', 'sm']) and xmlOp.text.startswith('ST'):
                  continue
               elif docOp == 'ax' and xmlOp.text == 'AX':
                  continue
               elif 'width' in xmlOp.attrib and re.search(r'r(eg)?(\d+/)*' + xmlOp.attrib['width'], docOp) is not None: continue
            elif xmlOp.attrib['type'] == 'mem':
               if docOp in ['mem', 'reg/mem', 'xmm2/mem', 'xmm3/mem', 'ymm3/mem', 'vm32x', 'm', '[r32/r64]/m', '[r8/r16]/m', 'g/m', 'gx/m', 'i/m', 'i/r/m',
                            'r/m', 's/m', 'ss/ds/es/fs/gs/m/r', 'x/m', 'xy/m', 'xyz/m', 'v']: continue
               if re.search(r'mem(\d+/)*' + xmlOp.attrib['width'], docOp) is not None: continue
               if (docOp in ['r8/m', 'r16/m', 'r32/m', 'r64/m']) and (re.search(r'\d+', docOp)[0] == xmlOp.attrib['width']): continue
            elif xmlOp.attrib['type'] in ['imm', 'relbr']:
               if docOp in ['imm', 'imm`', 'CL/Imm', 'xmm3/imm', 'i', 'i/cl', 'i/m', 'i/r', 'i/r/m', '3/4/i']: continue
               if re.search(r'imm(\d+/)*' + xmlOp.attrib['width'], docOp) is not None: continue
            invalid = True
            break

         if invalid:
            continue

         if xmlZeroing and 'z}' not in de.maskingOperand:
            continue
         if not xmlZeroing and de.maskingOperand in ['{k1}{z}', '{kz}']:
            continue

         if instrNode in xmlToDocDict:
            if (set(de.operands) != {None}) and (set(xmlToDocDict[instrNode].operands) == {None}):
               xmlFound = True
               xmlToDocDict[instrNode] = de
            elif (set(de.operands) == {None}) and (set(xmlToDocDict[instrNode].operands) != {None}):
               pass
            else:
               print('duplicate entry for ' + instrNode.attrib['string'] + ' found: ' + str(list(xmlToDocDict[instrNode])) + ', ' + str(list(de)))
         else:
            xmlFound = True
            xmlToDocDict[instrNode] = de

      if not xmlFound:
         print('no matching XML entry found for ' + str(de))

   print('Found data for ' + str(len(xmlToDocDict)) + ' instruction variants')

   for instrNode, de in xmlToDocDict.items():
      archNode = instrNode.find('./architecture[@name="{}"]'.format(args.arch))
      if archNode is None:
         archNode = ET.SubElement(instrNode, "architecture")
         archNode.attrib['name'] = args.arch

      docNode = ET.SubElement(archNode, "doc")
      if de.ops: docNode.attrib['uops'] = str(de.ops)
      if de.unit: docNode.attrib['ports'] = str(de.unit)
      if de.lat and de.lat != '-': docNode.attrib['latency'] = str(de.lat)
      if de.tp:
         try:
            if str(de.tp) == '0.33':
               docNode.attrib['TP'] = '3.00'
            else:
               docNode.attrib['TP'] = format(1/float(de.tp), '.2f')
         except ValueError:
            docNode.attrib['TP'] = de.tp

   with open(args.outputXML, "w") as f:
      rough_string = ET.tostring(root, 'utf-8')
      reparsed = minidom.parseString(rough_string)
      f.write('\n'.join([line for line in reparsed.toprettyxml(indent=' '*2).split('\n') if line.strip()]))


if __name__ == "__main__":
    main()
